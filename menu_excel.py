"""
Генерация Excel файла с меню на неделю.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_menu_excel(weekly_menu: dict, profile: dict, kbju: dict) -> str:
    """Создаёт Excel файл с меню на неделю. Возвращает путь к файлу."""

    wb = Workbook()
    wb.remove(wb.active)  # удаляем дефолтный лист

    # Цвета
    COLOR_HEADER    = "2E4057"  # тёмно-синий
    COLOR_DAY       = "048A81"  # зелёный
    COLOR_MEAL      = "E8F4F8"  # светло-голубой
    COLOR_TRAINING  = "FFF3CD"  # жёлтый для тренировочных дней
    COLOR_REST      = "F0F0F0"  # серый для дней отдыха
    COLOR_TOTAL     = "D4EDDA"  # зелёный для итогов
    WHITE           = "FFFFFF"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hfill(color): return PatternFill("solid", start_color=color, fgColor=color)
    def bold(size=11, color="000000"): return Font(name="Arial", bold=True, size=size, color=color)
    def normal(size=10): return Font(name="Arial", size=size)
    def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def left(): return Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── Лист 1: Сводное меню на неделю ───────────────────────────────────────
    ws = wb.create_sheet("Меню на неделю")
    ws.sheet_view.showGridLines = False

    # Заголовок
    ws.merge_cells("A1:H1")
    ws["A1"] = "🍽️ МЕНЮ ПИТАНИЯ НА НЕДЕЛЮ"
    ws["A1"].font = bold(16, WHITE)
    ws["A1"].fill = hfill(COLOR_HEADER)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 35

    # Подзаголовок с КБЖУ
    ws.merge_cells("A2:H2")
    ws["A2"] = (f"Цель: {kbju['target_calories']} ккал  |  "
                f"Б: {kbju['protein']}г  |  Ж: {kbju['fat']}г  |  У: {kbju['carbs']}г  |  "
                f"Фаза: {kbju['phase_label']}")
    ws["A2"].font = normal(10)
    ws["A2"].fill = hfill("E8EDF2")
    ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 20

    # Заголовки колонок
    headers = ["День", "Тип дня", "Приём пищи", "Продукты и граммовки",
               "Ккал", "Белки (г)", "Жиры (г)", "Углеводы (г)"]
    col_widths = [14, 14, 14, 45, 10, 12, 10, 14]

    row = 3
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = bold(10, WHITE)
        cell.fill = hfill(COLOR_DAY)
        cell.alignment = center()
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[row].height = 22
    row += 1

    # Данные по дням
    for day in weekly_menu.get("days", []):
        day_name  = day.get("day", "")
        day_type  = day.get("type", "")
        meals     = day.get("meals", [])
        day_color = COLOR_TRAINING if "тренировоч" in day_type.lower() else COLOR_REST
        day_rows_start = row

        for meal in meals:
            meal_name = meal.get("name", "")
            items = meal.get("items", [])
            items_text = "\n".join(
                f"• {i['product']}: {i['amount']} {i.get('unit','г')}"
                for i in items
            )
            cal = meal.get("calories", "")
            prot = meal.get("protein", "")
            fat  = meal.get("fat", "")
            carb = meal.get("carbs", "")

            ws.cell(row=row, column=1).fill = hfill(day_color)
            ws.cell(row=row, column=2).fill = hfill(day_color)
            ws.cell(row=row, column=3, value=meal_name).fill = hfill(COLOR_MEAL)
            ws.cell(row=row, column=3).font = bold(10)
            ws.cell(row=row, column=3).alignment = center()
            ws.cell(row=row, column=3).border = border
            ws.cell(row=row, column=4, value=items_text).alignment = left()
            ws.cell(row=row, column=4).font = normal(9)
            ws.cell(row=row, column=4).border = border
            ws.cell(row=row, column=5, value=cal).alignment = center()
            ws.cell(row=row, column=5).font = normal(10)
            ws.cell(row=row, column=5).border = border
            ws.cell(row=row, column=6, value=prot).alignment = center()
            ws.cell(row=row, column=6).font = normal(10)
            ws.cell(row=row, column=6).border = border
            ws.cell(row=row, column=7, value=fat).alignment = center()
            ws.cell(row=row, column=7).font = normal(10)
            ws.cell(row=row, column=7).border = border
            ws.cell(row=row, column=8, value=carb).alignment = center()
            ws.cell(row=row, column=8).font = normal(10)
            ws.cell(row=row, column=8).border = border
            ws.row_dimensions[row].height = max(15 * len(items), 25)
            row += 1

        # Объединяем ячейки дня и типа
        if row - day_rows_start > 1:
            ws.merge_cells(f"A{day_rows_start}:A{row-1}")
            ws.merge_cells(f"B{day_rows_start}:B{row-1}")

        ws.cell(row=day_rows_start, column=1, value=day_name).font = bold(11)
        ws.cell(row=day_rows_start, column=1).alignment = center()
        ws.cell(row=day_rows_start, column=1).fill = hfill(day_color)
        ws.cell(row=day_rows_start, column=1).border = border
        ws.cell(row=day_rows_start, column=2, value=day_type.capitalize()).font = normal(9)
        ws.cell(row=day_rows_start, column=2).alignment = center()
        ws.cell(row=day_rows_start, column=2).fill = hfill(day_color)
        ws.cell(row=day_rows_start, column=2).border = border

        # Итого за день
        total_cal  = day.get("total_calories", "")
        total_prot = day.get("total_protein", "")
        total_fat  = day.get("total_fat", "")
        total_carb = day.get("total_carbs", "")

        ws.merge_cells(f"A{row}:D{row}")
        ws.cell(row=row, column=1, value=f"Итого за день — {day_name}").font = bold(10)
        ws.cell(row=row, column=1).fill = hfill(COLOR_TOTAL)
        ws.cell(row=row, column=1).alignment = left()
        ws.cell(row=row, column=1).border = border
        for col, val in [(5, total_cal), (6, total_prot), (7, total_fat), (8, total_carb)]:
            ws.cell(row=row, column=col, value=val).font = bold(10)
            ws.cell(row=row, column=col).fill = hfill(COLOR_TOTAL)
            ws.cell(row=row, column=col).alignment = center()
            ws.cell(row=row, column=col).border = border
        ws.row_dimensions[row].height = 20
        row += 1

        # Пустая строка между днями
        ws.row_dimensions[row].height = 6
        row += 1

    # ── Лист 2: Список покупок ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Список покупок")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 15

    ws2.merge_cells("A1:C1")
    ws2["A1"] = "🛒 СПИСОК ПОКУПОК НА НЕДЕЛЮ"
    ws2["A1"].font = bold(14, WHITE)
    ws2["A1"].fill = hfill(COLOR_HEADER)
    ws2["A1"].alignment = center()
    ws2.row_dimensions[1].height = 30

    # Собираем продукты
    products = {}
    for day in weekly_menu.get("days", []):
        for meal in day.get("meals", []):
            for item in meal.get("items", []):
                key = item["product"]
                amount = item["amount"]
                unit = item.get("unit", "г")
                if key in products:
                    products[key]["amount"] += amount
                else:
                    products[key] = {"amount": amount, "unit": unit}

    # Категории
    categories = {
        "🥩 Белковые продукты": ["грудк", "бедр", "говяд", "индейк", "яйц", "творог",
                                   "тунец", "лосось", "минтай", "йогурт", "кефир"],
        "🍚 Крупы и углеводы":  ["овсян", "гречк", "рис", "макарон", "перловк",
                                   "пшен", "хлеб", "хлебц"],
        "🥦 Овощи и зелень":    ["брокколи", "огурц", "помидор", "перец", "морков",
                                   "капуст", "шпинат", "кабачок"],
        "🍎 Фрукты":            ["банан", "яблок", "апельсин"],
        "🥑 Жиры и орехи":      ["масло", "паст", "орех", "миндал", "авокадо"],
        "🥛 Молочное":          ["молок", "сыр", "молочн"],
    }

    row2 = 2
    # Заголовки
    for col, h in enumerate(["Продукт", "Количество", "Примечание"], 1):
        ws2.cell(row=row2, column=col, value=h).font = bold(10, WHITE)
        ws2.cell(row=row2, column=col).fill = hfill(COLOR_DAY)
        ws2.cell(row=row2, column=col).alignment = center()
        ws2.cell(row=row2, column=col).border = border
    row2 += 1

    categorized = set()
    for cat_name, keywords in categories.items():
        cat_items = {k: v for k, v in products.items()
                     if any(kw in k.lower() for kw in keywords)}
        if not cat_items:
            continue

        ws2.merge_cells(f"A{row2}:C{row2}")
        ws2.cell(row=row2, column=1, value=cat_name).font = bold(10, WHITE)
        ws2.cell(row=row2, column=1).fill = hfill(COLOR_DAY)
        ws2.cell(row=row2, column=1).alignment = left()
        ws2.row_dimensions[row2].height = 20
        row2 += 1

        for prod, data in cat_items.items():
            ws2.cell(row=row2, column=1, value=prod).font = normal(10)
            ws2.cell(row=row2, column=1).border = border
            ws2.cell(row=row2, column=2, value=f"{round(data['amount'])} {data['unit']}").font = normal(10)
            ws2.cell(row=row2, column=2).alignment = center()
            ws2.cell(row=row2, column=2).border = border
            ws2.cell(row=row2, column=3).border = border
            ws2.row_dimensions[row2].height = 18
            categorized.add(prod)
            row2 += 1

    # Прочее
    other = {k: v for k, v in products.items() if k not in categorized}
    if other:
        ws2.merge_cells(f"A{row2}:C{row2}")
        ws2.cell(row=row2, column=1, value="🧂 Прочее").font = bold(10, WHITE)
        ws2.cell(row=row2, column=1).fill = hfill(COLOR_DAY)
        row2 += 1
        for prod, data in other.items():
            ws2.cell(row=row2, column=1, value=prod).font = normal(10)
            ws2.cell(row=row2, column=1).border = border
            ws2.cell(row=row2, column=2, value=f"{round(data['amount'])} {data['unit']}").font = normal(10)
            ws2.cell(row=row2, column=2).alignment = center()
            ws2.cell(row=row2, column=2).border = border
            ws2.cell(row=row2, column=3).border = border
            row2 += 1

    path = "menu_week.xlsx"
    wb.save(path)
    return path
